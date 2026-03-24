import sys
import json
import base64
import re
import hashlib
import platform
import uuid
from pathlib import Path
from datetime import datetime
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QListWidget, QLabel, QMessageBox, QFileDialog,
    QProgressBar, QLineEdit, QDialog, QTextEdit, QListWidgetItem
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font as ExcelFont, Alignment, Border, Side

# --- APIキー安全保存のためのモジュール ---
# 優先順位: keyring（OS資格情報ストア） > Fernet暗号化 > 保存不可
_USE_KEYRING = False
_USE_FERNET = False

try:
    import keyring
    # keyringが実際に動作するかテスト（バックエンドがない環境対策）
    keyring.get_keyring()
    _USE_KEYRING = True
except Exception:
    pass

if not _USE_KEYRING:
    try:
        from cryptography.fernet import Fernet
        _USE_FERNET = True
    except ImportError:
        pass

# アプリケーション識別子（keyring用）
_KEYRING_SERVICE = "evidence_schedule_tool"
_KEYRING_USERNAME = "claude_api_key"
# Fernet暗号化用の設定ファイルパス
_CONFIG_FILE = Path.home() / ".evidence_schedule_config.json"


def _get_machine_key() -> bytes:
    """マシン固有の情報からFernet暗号化キーを生成する。
    
    MACアドレス + ユーザー名 + マシン名からSHA256ハッシュを生成し、
    Fernetキーとして使用する。これにより設定ファイルを他のPCに
    コピーしても復号できない。
    """
    machine_info = f"{uuid.getnode()}-{platform.node()}-{Path.home()}"
    key_hash = hashlib.sha256(machine_info.encode()).digest()
    return base64.urlsafe_b64encode(key_hash)


def save_api_key_secure(api_key: str) -> None:
    """APIキーを安全に保存する。
    
    - keyringが利用可能: OS標準の資格情報ストア（Windows Credential Manager等）に保存
    - cryptographyが利用可能: マシン固有キーでFernet暗号化してファイルに保存
    - いずれも不可: 保存せず例外を発生
    """
    if _USE_KEYRING:
        keyring.set_password(_KEYRING_SERVICE, _KEYRING_USERNAME, api_key)
    elif _USE_FERNET:
        fernet = Fernet(_get_machine_key())
        encrypted = fernet.encrypt(api_key.encode()).decode()
        with open(_CONFIG_FILE, 'w') as f:
            json.dump({'api_key_encrypted': encrypted}, f)
    else:
        raise RuntimeError(
            "APIキーを安全に保存するためのライブラリがインストールされていません。\n"
            "以下のいずれかをインストールしてください：\n"
            "  pip install keyring\n"
            "  pip install cryptography"
        )


def load_api_key_secure() -> str:
    """安全に保存されたAPIキーを読み込む。"""
    if _USE_KEYRING:
        key = keyring.get_password(_KEYRING_SERVICE, _KEYRING_USERNAME)
        return key if key else ''
    elif _USE_FERNET:
        if _CONFIG_FILE.exists():
            try:
                with open(_CONFIG_FILE, 'r') as f:
                    config = json.load(f)
                encrypted = config.get('api_key_encrypted', '')
                if encrypted:
                    fernet = Fernet(_get_machine_key())
                    return fernet.decrypt(encrypted.encode()).decode()
            except Exception:
                # 復号失敗（別マシンのファイル、破損等）→ファイルを削除
                _CONFIG_FILE.unlink(missing_ok=True)
        return ''
    else:
        # 旧バージョンの平文設定ファイルがあれば読み込みを試みる（移行用）
        if _CONFIG_FILE.exists():
            try:
                with open(_CONFIG_FILE, 'r') as f:
                    config = json.load(f)
                return config.get('api_key', '')
            except Exception:
                pass
        return ''


def delete_api_key_secure() -> None:
    """保存されたAPIキーを削除する。"""
    if _USE_KEYRING:
        try:
            keyring.delete_password(_KEYRING_SERVICE, _KEYRING_USERNAME)
        except keyring.errors.PasswordDeleteError:
            pass  # 既に存在しない場合
    # ファイルがあれば削除（keyring使用時でも旧ファイルがあれば消す）
    if _CONFIG_FILE.exists():
        _CONFIG_FILE.unlink()


def get_storage_method_description() -> str:
    """現在のAPIキー保存方式の説明文を返す。"""
    if _USE_KEYRING:
        if platform.system() == 'Windows':
            return "入力したAPIキーはWindows資格情報マネージャーに暗号化して保存されます。"
        elif platform.system() == 'Darwin':
            return "入力したAPIキーはmacOSキーチェーンに暗号化して保存されます。"
        else:
            return "入力したAPIキーはOSの資格情報ストアに暗号化して保存されます。"
    elif _USE_FERNET:
        return "入力したAPIキーはマシン固有の鍵で暗号化してファイルに保存されます。"
    else:
        return (
            "⚠ APIキーの安全な保存に必要なライブラリが見つかりません。\n"
            "APIキーは今回の起動中のみ有効で、終了時に破棄されます。\n"
            "永続保存するには keyring または cryptography をインストールしてください。"
        )


def to_wareki(date_str: str) -> str:
    """西暦日付文字列を和暦に変換する。
    
    入力例: "2024年3月15日" → "R6.3.15"
    対応元号: 令和(R), 平成(H), 昭和(S)
    変換できない場合は入力をそのまま返す。
    """
    if not date_str:
        return date_str

    # "YYYY年M月D日" 形式を解析
    m = re.match(r'(\d{4})年(\d{1,2})月(\d{1,2})日', date_str.strip())
    if not m:
        return date_str

    year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))

    # 元号テーブル（新しい順に評価）
    # (開始年, 開始月, 開始日, 略称, 元年オフセット)
    ERA_TABLE = [
        (2019, 5, 1,  'R', 2019),  # 令和: 2019-05-01〜
        (1989, 1, 8,  'H', 1989),  # 平成: 1989-01-08〜2019-04-30
        (1926, 12, 25, 'S', 1926), # 昭和: 1926-12-25〜1989-01-07
    ]

    date_int = year * 10000 + month * 100 + day

    for start_y, start_m, start_d, abbr, base in ERA_TABLE:
        start_int = start_y * 10000 + start_m * 100 + start_d
        if date_int >= start_int:
            era_year = year - base + 1
            return f"{abbr}{era_year}.{month}.{day}"

    # 昭和より前（大正・明治など）はそのまま返す
    return date_str


def format_evidence_short(evidence_number: str) -> str:
    """証拠番号を短縮表示形式に変換する。

    入力例:
      "甲第1号証"     → "甲1"
      "甲第01号証"    → "甲1"
      "甲第1号証の1"  → "甲1の1"
      "第5号証"       → "5"
    """
    # パターン1: 甲第1号証の2 など（枝番あり）
    m = re.match(r'(甲|乙|丙|丁)?第0*(\d+)号証の0*(\d+)', evidence_number)
    if m:
        prefix = m.group(1) or ''
        main = m.group(2)
        branch = m.group(3)
        return f"{prefix}{main}の{branch}"

    # パターン2: 甲第1号証 など（枝番なし）
    m = re.match(r'(甲|乙|丙|丁)?第0*(\d+)号証', evidence_number)
    if m:
        prefix = m.group(1) or ''
        main = m.group(2)
        return f"{prefix}{main}"

    # マッチしない場合はそのまま
    return evidence_number


class TextViewerDialog(QDialog):
    """テキスト全文表示用の子ダイアログ"""

    def __init__(self, parent, title: str, content: str):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(620, 500)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)

        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText(content)
        text_edit.setFont(QFont("Yu Gothic UI", 9))
        text_edit.moveCursor(text_edit.textCursor().MoveOperation.Start)
        layout.addWidget(text_edit)

        close_btn = QPushButton("閉じる")
        close_btn.setFixedWidth(100)
        close_btn.clicked.connect(self.accept)
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)


class AboutDialog(QDialog):
    """カスタムAboutダイアログ（操作説明書・README・ライセンス情報へのリンク付き）"""

    # --- 埋め込みテキスト定数 ---
    # MSIX / Microsoft Store 配布前提で改訂済み

    README_TEXT = (
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "証拠説明書作成支援ツール\n"
        "README\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "この度は証拠説明書作成支援ツールをご利用いただき、\n"
        "誠にありがとうございます。\n\n"
        "本ツールは、訴訟・紛争案件における証拠説明書の作成業務を\n"
        "効率化するために開発された専用ツールです。\n"
        "PDFファイルの1ページ目をAI（Claude Opus 4.6）が読み取り、\n"
        "標題・作成者・作成年月日・立証趣旨を自動抽出して\n"
        "Excel形式の証拠説明書案を生成します。\n"
        "また、生成したExcelの内容をもとに証拠PDFのファイル名末尾に\n"
        "証拠標題を付加するリネーム機能も備えています。\n\n\n"
        "■ 動作環境\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "OS：Windows 10 / 11（64bit）\n"
        "メモリ：8GB以上推奨\n"
        "ストレージ：500MB以上の空き容量\n"
        "インターネット接続：必須（Claude API通信のため）\n\n\n"
        "■ 事前準備：Claude APIキーの取得\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "本ツールの使用にはAnthropic社のClaude APIキーが必要です。\n"
        "APIの利用には従量課金の料金が発生します。\n\n"
        "【APIキーの取得方法】\n"
        "1. https://console.anthropic.com/ にアクセス\n"
        "2. アカウントを作成またはログイン\n"
        "3. Settings → API Keys で新しいキーを作成\n"
        "4. 表示されたキー（sk-ant-...）をコピー\n\n"
        "【料金目安】\n"
        "  証拠100件：約$1\n"
        "  証拠1000件：約$10\n"
        "  ※Claude Opus 4.6使用時の概算です\n\n\n"
        "■ 起動方法\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Microsoft Storeからインストール後、スタートメニューから起動してください。\n"
        "初回起動時にAPIキーの入力が求められます。\n\n\n"
        "■ ファイル名の命名規則\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "証拠番号の自動割り当てに対応するため、\n"
        "PDFファイル名は以下の形式を推奨します。\n\n"
        "  甲01.pdf  甲02.pdf  甲03の1.pdf  甲03の2.pdf\n"
        "  乙01.pdf  乙02.pdf\n\n"
        "ファイル名に甲・乙・丙・丁と数字が含まれていれば、\n"
        "自動的に「甲第1号証」「乙第2号証」等に変換されます。\n"
        "枝番は「の」で区切って表記します（例：甲03の1）。\n\n"
        "数字のみのファイル名（1.pdf、2.pdf等）や\n"
        "上記パターンに該当しないファイルは自動採番されます。\n\n\n"
        "■ 出力されるExcelの項目\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "証拠種別 ｜ 番号 ｜ 標題 ｜ 原本・写しの別 ｜ 作成者 ｜ 作成年月日 ｜ 立証趣旨\n\n"
        "  証拠種別：甲・乙・丙・丁（ファイル名から自動判定）\n"
        "  番号：証拠番号（ファイル名から自動判定）\n"
        "  標題：AIが文書から自動抽出\n"
        "  原本・写しの別：「写し」（固定）\n"
        "  作成者：AIが文書から自動抽出\n"
        "  作成年月日：AIが文書から自動抽出（和暦表記）\n"
        "  立証趣旨：AIが文書内容から自動推定\n\n"
        "※AIの抽出結果は必ず確認・修正してください。\n"
        "  出力はあくまで下書きであり、最終確認は人間が行う必要があります。\n"
        "※H列にファイルパスが自動記録されます（非表示）。リネーム機能で使用します。\n\n\n"
        "■ PDFリネーム機能\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "証拠説明書案の内容確認・修正後に、PDFのファイル名末尾に\n"
        "証拠標題を付加したコピーを作成します（元ファイルは変更されません）。\n\n"
        "【使い方】\n"
        "1. 証拠説明書案のExcel（またはCSV）をリネームエリアにドロップ\n"
        "2. 「リネーム実行」をクリック\n"
        "3. 対応するフォルダの隣に「_リネーム済」フォルダが作成され、\n"
        "   リネームされたPDFが保存される\n\n"
        "【リネーム後のファイル名の例】\n"
        "  甲01_業務委託契約書.pdf\n"
        "  甲02_請求書.pdf\n"
        "  甲03の1_覚書.pdf\n\n\n"
        "■ プライバシーとデータ保護について\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "【送信されるデータ】\n"
        "本ツールは各PDFファイルの「1ページ目のみ」を画像として\n"
        "Claude APIに送信します。2ページ目以降は一切送信されません。\n\n"
        "【データの取り扱い（Claude API）】\n"
        "・送信データは学習に使用されません\n"
        "・不正利用検出のため30日間保持され、その後自動削除されます\n"
        "・詳細はAnthropic社のプライバシーポリシーをご確認ください\n\n"
        "【ユーザーの責任】\n"
        "1ページ目に個人情報や機密情報が含まれる場合は、\n"
        "事前に墨消し処理を行うなど、適切な対策を講じてください。\n"
        "本ツールの使用により生じた情報漏洩等の責任は、\n"
        "ユーザーが負うものとします。\n\n\n"
        "■ よくある質問\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Q. インターネット接続は必要ですか？\n"
        "A. はい。Claude APIとの通信のため、常時インターネット接続が必要です。\n\n"
        "Q. API料金はどのくらいかかりますか？\n"
        "A. 証拠100件で約$1が目安です。\n\n"
        "Q. 元のPDFファイルが変更されることはありますか？\n"
        "A. ありません。リネーム機能もコピーを作成するため元ファイルは変更されません。\n\n"
        "Q. AIの抽出結果はどの程度正確ですか？\n"
        "A. Claude Opus 4.6を使用しており高い精度で抽出できますが、\n"
        "   完全な正確性は保証されません。出力結果は必ず確認・修正してください。\n\n\n"
        "■ 免責事項\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "本ソフトウェアの使用により生じたいかなる損害についても、\n"
        "開発者は一切の責任を負いかねます。\n"
        "AIによる抽出結果は参考情報であり、正確性を保証するものではありません。\n\n\n"
        "■ 著作権とライセンス\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "開発・販売：Lean Tech Library\n\n"
        "本ソフトウェアはAGPL-3.0ライセンスの下で配布されています。\n"
        "再配布の際はライセンス条件に従ってください。\n\n"
        "ソースコード：\n"
        "https://github.com/leantechlibrary-coder/LTL-Evidence-Schedule-Tool\n"
    )

    MANUAL_TEXT = (
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "証拠説明書作成支援ツール\n"
        "操作説明書\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n\n"
        "■ 目次\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "  1. ツールの概要\n"
        "  2. 事前準備（APIキーの取得）\n"
        "  3. 初回起動とAPIキーの設定\n"
        "  4. 基本操作の流れ\n"
        "  5. 各ボタンの説明\n"
        "  6. ファイル名と証拠番号の対応\n"
        "  7. 出力されるExcelファイルの内容\n"
        "  8. PDFリネーム機能\n"
        "  9. APIキーの管理\n"
        "  10. プライバシーとデータ保護について\n"
        "  11. よくある質問（FAQ）\n"
        "  12. トラブルシューティング\n"
        "  13. 注意事項\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "1. ツールの概要\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "本ツールは、訴訟・紛争案件における証拠説明書の作成を\n"
        "AIで自動化するツールです。\n\n"
        "【できること】\n"
        "・PDFファイルの1ページ目をAIが読み取る\n"
        "・標題（文書タイトル）を自動抽出\n"
        "・作成者（会社名・個人名）を自動抽出\n"
        "・作成年月日を自動抽出（和暦表記）\n"
        "・立証趣旨を自動推定\n"
        "・Excel形式（.xlsx）の証拠説明書案を自動生成\n"
        "・証拠説明書案をもとに証拠PDFのファイル名をリネーム（任意）\n\n"
        "【使用するAI】\n"
        "Claude Opus 4.6（Anthropic社の最高精度モデル）\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "2. 事前準備（APIキーの取得）\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "本ツールを使用するには、Anthropic社のClaude APIキーが必要です。\n"
        "APIの利用には従量課金の料金が発生します。\n\n"
        "【手順】\n\n"
        "(1) Anthropic社のコンソールにアクセス\n"
        "    https://console.anthropic.com/\n\n"
        "(2) アカウントを作成（初回のみ）\n"
        "    - メールアドレスで登録\n"
        "    - 支払い方法（クレジットカード）を登録\n\n"
        "(3) APIキーを作成\n"
        "    - 画面左の Settings → API Keys を選択\n"
        "    - 「Create Key」をクリック\n"
        "    - キーの名前を入力（例：「証拠説明書ツール」）\n"
        "    - 表示されたキーをコピー\n\n"
        "    ※キーは「sk-ant-」で始まる文字列です\n"
        "    ※キーは作成時に一度だけ表示されます。必ずコピーしてください\n\n"
        "【料金目安】\n"
        "    証拠10件：約$0.1\n"
        "    証拠100件：約$1\n"
        "    証拠1000件：約$10\n\n"
        "    ※上記はClaude Opus 4.6使用時の概算です\n"
        "    ※実際の料金はAnthropic社の価格設定によります\n"
        "    ※Anthropic社のコンソールでUsageから利用状況を確認できます\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "3. 初回起動とAPIキーの設定\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "(1) スタートメニューからツールを起動\n\n"
        "(2) 「APIキー設定」ダイアログが自動的に表示されます\n\n"
        "(3) 事前にコピーしたAPIキーを入力欄に貼り付け\n"
        "    - 入力欄の右にある「表示」ボタンで入力内容を確認できます\n\n"
        "(4) 「OK」をクリック\n\n"
        "(5) 「APIキーを保存しました」と表示されれば設定完了\n\n"
        "    ※APIキーはWindows資格情報マネージャーに暗号化して保存されます\n"
        "    ※次回以降の起動時にはAPIキーの入力は不要です\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "4. 基本操作の流れ\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "【ステップ1】PDFファイルの読み込み\n\n"
        "  「フォルダを開く」ボタンをクリック\n"
        "  → PDFファイルが入ったフォルダを選択\n"
        "  → フォルダ内のPDFファイルが一覧に表示されます\n\n"
        "  ファイルはファイル名に基づいて自動的にソートされ、\n"
        "  証拠番号（甲第○号証 等）が割り当てられます。\n\n"
        "【ステップ2】ファイル一覧の確認\n\n"
        "  画面中央のリストに、以下の形式で表示されます：\n"
        "    甲1 - 甲01.pdf\n"
        "    甲2 - 甲02.pdf\n"
        "    甲3の1 - 甲03の1.pdf\n"
        "    ...\n\n"
        "  証拠番号の割り当てが正しいことを確認してください。\n\n"
        "【ステップ3】証拠説明書案の生成\n\n"
        "  「生成実行」ボタン（緑色）をクリック\n"
        "  → API料金の確認ダイアログが表示されます\n"
        "  → 「はい」をクリック\n\n"
        "  処理中はプログレスバーで進捗が表示されます。\n"
        "  処理が完了すると、読み込んだPDFと同じフォルダに\n"
        "  「フォルダ名_証拠説明書案.xlsx」が自動保存されます。\n\n"
        "【ステップ4】出力結果の確認・修正\n\n"
        "  Excelファイルを開き、AIが抽出した内容を確認してください。\n"
        "  必要に応じて標題・作成者・作成年月日・立証趣旨を修正します。\n\n"
        "【ステップ5】PDFリネーム（任意）\n\n"
        "  証拠説明書の内容が確定したら、リネーム機能を使って\n"
        "  PDFのファイル名末尾に証拠標題を付加できます。\n"
        "  詳しくは「8. PDFリネーム機能」をご覧ください。\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "5. 各ボタンの説明\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "【フォルダを開く】\n"
        "  フォルダを選択し、中のPDFファイルをまとめて読み込みます。\n\n"
        "【ファイルを追加】\n"
        "  個別のPDFファイルを選択して、既存のリストに追加します。\n\n"
        "【クリア】\n"
        "  読み込んだファイルの一覧をすべて消去します。\n\n"
        "【APIキー設定】\n"
        "  APIキーの新規入力・変更を行います。\n\n"
        "【APIキー削除】\n"
        "  保存されているAPIキーを完全に削除します。\n\n"
        "【生成実行】（緑色ボタン）\n"
        "  証拠説明書案の自動生成を開始します。\n\n"
        "【ファイルをクリア】（リネームエリア）\n"
        "  ドロップしたExcel/CSVファイルをクリアします。\n\n"
        "【リネーム実行】（青色ボタン）\n"
        "  ドロップしたExcel/CSVの内容をもとにPDFをリネームします。\n\n"
        "【About】（画面右上）\n"
        "  バージョン情報、操作説明書、README、ライセンス情報を表示します。\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "6. ファイル名と証拠番号の対応\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "本ツールはPDFのファイル名から証拠種別と番号を自動判定します。\n\n"
        "【対応するファイル名の例】\n\n"
        "  ファイル名          →  証拠番号\n"
        "  ─────────────────────────────────\n"
        "  甲01.pdf            →  甲第1号証\n"
        "  甲02.pdf            →  甲第2号証\n"
        "  甲03の1.pdf         →  甲第3号証の1\n"
        "  甲03の2.pdf         →  甲第3号証の2\n"
        "  乙01.pdf            →  乙第1号証\n\n"
        "【対応する証拠種別】\n"
        "  甲・乙・丙・丁\n\n"
        "【枝番の表記】\n"
        "  「の」で区切って表記します（例：甲03の1.pdf）。\n"
        "  ハイフン区切り（甲03-1.pdf）も使用できます。\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "7. 出力されるExcelファイルの内容\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "  A列：証拠種別（甲・乙・丙・丁）\n"
        "  B列：番号（1, 2, 3の1 等）\n"
        "  C列：標題（AIが抽出した文書タイトル）\n"
        "  D列：原本・写しの別（「写し」固定）\n"
        "  E列：作成者（AIが抽出した作成者名）\n"
        "  F列：作成年月日（AIが抽出した日付・和暦表記）\n"
        "  G列：立証趣旨（AIが推定した立証趣旨）\n"
        "  H列：ファイルパス（非表示・リネーム機能で使用）\n\n"
        "【重要】\n"
        "  AIの抽出・推定結果はあくまで下書きです。\n"
        "  出力後に必ず内容を確認し、必要に応じて修正してください。\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "8. PDFリネーム機能\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "証拠説明書案の内容が確定した後、PDFのファイル名末尾に\n"
        "証拠標題を付加したコピーを作成します。\n"
        "元のPDFファイルは変更されません。\n\n"
        "【使い方】\n\n"
        "(1) 証拠説明書案のExcel（.xlsx）またはCSVを\n"
        "    画面下部のドロップエリアにドラッグ＆ドロップ\n\n"
        "(2) 「リネーム実行」ボタン（青色）をクリック\n\n"
        "(3) 完了ダイアログが表示され、出力先フォルダが自動で開く\n\n"
        "【出力先】\n"
        "  元のPDFフォルダの隣に「_リネーム済」フォルダが作成されます。\n\n"
        "【リネーム後のファイル名】\n"
        "  例：\n"
        "    甲01_業務委託契約書.pdf\n"
        "    甲02_請求書.pdf\n"
        "    甲03の1_覚書.pdf\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "9. APIキーの管理\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "【保存方法】\n"
        "APIキーはWindows資格情報マネージャーに暗号化して保存されます。\n"
        "平文のファイルとして保存されることはありません。\n\n"
        "【APIキーの変更】\n"
        "「APIキー設定」ボタンから新しいキーを入力すると上書きされます。\n\n"
        "【APIキーの削除】\n"
        "「APIキー削除」ボタンをクリックすると完全に削除されます。\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "10. プライバシーとデータ保護について\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "【送信されるデータ】\n"
        "本ツールは各PDFファイルの「1ページ目のみ」を画像として\n"
        "Claude APIに送信します。2ページ目以降は一切送信されません。\n\n"
        "【データの取り扱い（Claude API）】\n"
        "・送信データは学習に使用されません\n"
        "・不正利用検出のため30日間保持され、その後自動削除されます\n"
        "・詳細はAnthropic社のプライバシーポリシーをご確認ください\n\n"
        "【ユーザーの責任】\n"
        "1ページ目に個人情報や機密情報が含まれる場合は、\n"
        "事前に墨消し処理を行うなど、適切な対策を講じてください。\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "11. よくある質問（FAQ）\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "Q. インターネット接続は必要ですか？\n"
        "A. はい。Claude APIとの通信のため必要です。\n\n"
        "Q. API料金はどのくらいかかりますか？\n"
        "A. 証拠100件で約$1が目安です。\n\n"
        "Q. 元のPDFファイルが変更されることはありますか？\n"
        "A. ありません。リネーム機能もコピーを作成するため安全です。\n\n"
        "Q. 「原本・写しの別」が常に「写し」になっています。\n"
        "A. 仕様です。PDFは電子データであるため一律「写し」としています。\n"
        "   原本の場合は、出力後にExcel上で修正してください。\n\n"
        "Q. 立証趣旨が適切でない場合があります。\n"
        "A. AIは文書の1ページ目のみから推定しています。\n"
        "   案件の文脈に応じて人間が修正してください。\n\n"
        "Q. パスワード保護されたPDFは使えますか？\n"
        "A. 対応していません。パスワードを解除してからご使用ください。\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "12. トラブルシューティング\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "【ツールが起動しない】\n"
        "・Windows 10 / 11（64bit）であることを確認してください\n"
        "・Windowsを最新の状態に更新してください\n\n"
        "【APIキー関連のエラー】\n"
        "・APIキーが正しくコピーされているか確認してください\n"
        "  （先頭の sk-ant- を含む全体をコピー）\n"
        "・Anthropic社のコンソールでAPIキーが有効か確認してください\n"
        "・API利用料金の残高が十分か確認してください\n\n"
        "【処理中にエラーが表示される】\n"
        "・インターネット接続を確認してください\n"
        "・PDFファイルが破損していないか確認してください\n"
        "・ファイル名に特殊文字が含まれていないか確認してください\n\n"
        "【処理が遅い】\n"
        "・1件あたり数秒～10秒程度かかります（API通信のため）\n"
        "・100件で数分程度が目安です\n\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "13. 注意事項\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "・本ソフトウェアは「現状有姿」(AS IS) で提供されます\n"
        "・本ソフトウェアの使用により生じたいかなる損害についても、\n"
        "  開発者は一切の責任を負いかねます\n"
        "・AIによる抽出・推定結果は参考情報です\n"
        "・証拠説明書の最終的な内容は人間が確認・修正してください\n"
        "・各PDFの1ページ目がインターネット経由で送信されます\n"
        "・機密性の高い情報は事前に墨消し処理を行ってください\n"
    )

    LICENSE_TEXT = (
        "================================================================================\n"
        "THIRD-PARTY SOFTWARE LICENSES\n"
        "証拠説明書作成支援ツール\n"
        "================================================================================\n\n"
        "本ソフトウェアは、以下のオープンソースソフトウェアを使用しています。\n"
        "各ソフトウェアのライセンス条項に従い、ライセンス情報を記載します。\n\n\n"
        "================================================================================\n"
        "1. PyMuPDF (fitz)\n"
        "================================================================================\n\n"
        "License: GNU Affero General Public License v3.0 (AGPL-3.0)\n"
        "Copyright: Artifex Software, Inc.\n"
        "Website: https://github.com/pymupdf/PyMuPDF\n\n"
        "ライセンス全文：https://www.gnu.org/licenses/agpl-3.0.txt\n\n\n"
        "================================================================================\n"
        "2. PyQt6\n"
        "================================================================================\n\n"
        "License: GNU General Public License v3.0 (GPL-3.0)\n"
        "Copyright: Riverbank Computing Limited\n"
        "Website: https://www.riverbankcomputing.com/software/pyqt/\n\n"
        "ライセンス全文：https://www.gnu.org/licenses/gpl-3.0.txt\n\n\n"
        "================================================================================\n"
        "3. openpyxl\n"
        "================================================================================\n\n"
        "License: MIT License\n"
        "Copyright: Eric Gazoni, Charlie Clark\n"
        "Website: https://openpyxl.readthedocs.io/\n\n\n"
        "================================================================================\n"
        "4. Requests\n"
        "================================================================================\n\n"
        "License: Apache License 2.0\n"
        "Copyright: Kenneth Reitz\n"
        "Website: https://requests.readthedocs.io/\n\n\n"
        "================================================================================\n"
        "5. keyring\n"
        "================================================================================\n\n"
        "License: MIT License / Python Software Foundation License\n"
        "Copyright: Jason R. Coombs\n"
        "Website: https://github.com/jaraco/keyring\n\n\n"
        "================================================================================\n"
        "6. cryptography\n"
        "================================================================================\n\n"
        "License: Apache License 2.0 / BSD 3-Clause License\n"
        "Copyright: The Python Cryptographic Authority developers\n"
        "Website: https://github.com/pyca/cryptography\n\n\n"
        "================================================================================\n"
        "7. Python\n"
        "================================================================================\n\n"
        "License: Python Software Foundation License (PSF)\n"
        "Copyright: Python Software Foundation\n"
        "Website: https://www.python.org/\n\n"
        "ライセンス全文：https://docs.python.org/3/license.html\n\n\n"
        "================================================================================\n"
        "本ソフトウェアのライセンス\n"
        "================================================================================\n\n"
        "本ソフトウェア（証拠説明書作成支援ツール）は、\n"
        "GNU Affero General Public License v3.0 (AGPL-3.0) の下で配布されます。\n"
        "再配布の際はライセンス条件に従ってください。\n\n"
        "ソースコード：\n"
        "https://github.com/leantechlibrary-coder/LTL-Evidence-Schedule-Tool\n\n\n"
        "================================================================================\n"
        "免責事項\n"
        "================================================================================\n\n"
        "本ソフトウェアは「現状有姿」(AS IS) で提供され、いかなる保証もありません。\n"
        "本ソフトウェアの使用により生じたいかなる損害についても、開発者は\n"
        "一切の責任を負いません。\n"
    )

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("このソフトについて")
        self.resize(520, 480)
        self.setMinimumSize(400, 350)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        # --- タイトル ---
        title_label = QLabel("証拠説明書作成支援ツール v1.0")
        title_label.setFont(QFont("Yu Gothic UI", 12, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)

        # --- 本文（スクロール可能） ---
        about_text = QTextEdit()
        about_text.setReadOnly(True)
        about_text.setFont(QFont("Yu Gothic UI", 9))
        about_text.setPlainText(
            "【動作環境】\n"
            "Windows 10 / 11 (64bit)\n\n"
            "【使用AI】\n"
            "Claude Opus 4.6（最高精度モデル）\n\n"
            "【重要】プライバシーについて\n"
            "本ツールは各PDFの1ページ目のみをClaude APIに送信します。\n"
            "2ページ目以降は送信されません。\n"
            "送信データは学習に使用されず、30日後に自動削除されます。\n\n"
            "ただし、1ページ目に個人情報や機密情報が含まれる場合は、\n"
            "事前に墨消し処理を行うなど、適切な対策を講じてください。\n"
            "本ツールの使用により生じた情報漏洩等の責任は\n"
            "ユーザーが負うものとします。\n\n"
            "【APIキーの保存】\n"
            f"{get_storage_method_description()}\n\n"
            "【免責事項】\n"
            "本ソフトウェアは「現状有姿」(AS IS) で提供されます。\n"
            "本ソフトウェアの使用により生じたいかなる損害についても、\n"
            "開発者は一切の責任を負いません。\n"
            "重要なファイルは必ずバックアップを取ってからご使用ください。\n\n"
            "【開発・販売】\n"
            "Lean Tech Library\n\n"
            "ご使用前に操作説明書・READMEをご確認ください。"
        )
        layout.addWidget(about_text)

        # --- 詳細情報リンクボタン群 ---
        link_layout = QHBoxLayout()
        link_layout.setSpacing(8)

        manual_btn = QPushButton("操作説明書")
        manual_btn.setToolTip("操作説明書を表示します")
        manual_btn.clicked.connect(self._show_manual)

        readme_btn = QPushButton("README")
        readme_btn.setToolTip("READMEを表示します")
        readme_btn.clicked.connect(self._show_readme)

        license_btn = QPushButton("ライセンス情報")
        license_btn.setToolTip("サードパーティライセンス情報を表示します")
        license_btn.clicked.connect(self._show_licenses)

        link_layout.addWidget(manual_btn)
        link_layout.addWidget(readme_btn)
        link_layout.addWidget(license_btn)
        layout.addLayout(link_layout)

        # --- 閉じるボタン ---
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_btn = QPushButton("閉じる")
        close_btn.setFixedWidth(100)
        close_btn.clicked.connect(self.accept)
        close_layout.addWidget(close_btn)
        close_layout.addStretch()
        layout.addLayout(close_layout)

    def _show_manual(self):
        dlg = TextViewerDialog(self, "操作説明書", self.MANUAL_TEXT)
        dlg.exec()

    def _show_readme(self):
        dlg = TextViewerDialog(self, "README", self.README_TEXT)
        dlg.exec()

    def _show_licenses(self):
        dlg = TextViewerDialog(self, "ライセンス情報", self.LICENSE_TEXT)
        dlg.exec()


def show_about_dialog():
    """Aboutダイアログを表示"""
    dlg = AboutDialog()
    dlg.exec()


class APIKeyDialog(QDialog):
    """APIキー設定ダイアログ"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("APIキー設定")
        self.setModal(True)
        self.setMinimumWidth(500)
        self._real_text = ""  # 実際の入力テキストを保持
        self._updating = False  # 再帰防止フラグ
        
        layout = QVBoxLayout()
        
        # 説明（保存方式に応じて動的に変更）
        storage_desc = get_storage_method_description()
        info_label = QLabel(
            "Claude APIキーを入力してください。\n"
            "APIキーは https://console.anthropic.com/settings/keys で取得できます。\n"
            f"{storage_desc}\n\n"
            "※本ツールはClaude Opus 4.6を使用します（高精度）"
        )
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # APIキー入力
        # EchoMode.Password を使わず、独自マスクで表示する。
        # これによりWindowsがパスワード欄と認識して自動入力する問題を回避する。
        api_key_layout = QHBoxLayout()
        api_key_label = QLabel("APIキー:")
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("ここにAPIキーを貼り付けてください")
        # EchoModeはNormalのまま（Passwordにしない）
        self.api_key_input.setEchoMode(QLineEdit.EchoMode.Normal)
        # IME・オートコンプリートを無効化
        self.api_key_input.setAttribute(Qt.WidgetAttribute.WA_InputMethodEnabled, False)
        # テキスト変更時に独自マスク処理を実行
        self.api_key_input.textChanged.connect(self._on_text_changed)
        api_key_layout.addWidget(api_key_label)
        api_key_layout.addWidget(self.api_key_input)
        
        # 表示/非表示切替ボタン
        self.toggle_visibility_btn = QPushButton("表示")
        self.toggle_visibility_btn.setFixedWidth(50)
        self.toggle_visibility_btn.setCheckable(True)
        self.toggle_visibility_btn.clicked.connect(self._toggle_visibility)
        api_key_layout.addWidget(self.toggle_visibility_btn)
        
        layout.addLayout(api_key_layout)
        
        # 料金情報
        cost_info = QLabel(
            "料金目安：\n"
            "• 証拠100件：約$1\n"
            "• 証拠1000件：約$10"
        )
        cost_info.setStyleSheet("""
            QLabel {
                background-color: #F0F8FF;
                border: 1px solid #4682B4;
                border-radius: 5px;
                padding: 10px;
                color: #333;
            }
        """)
        layout.addWidget(cost_info)
        
        # ボタン
        button_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = QPushButton("キャンセル")
        self.cancel_button.clicked.connect(self.reject)
        button_layout.addStretch()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def _on_text_changed(self, displayed_text):
        """独自マスク処理：入力された文字を受け取り、実テキストを更新して●で表示する"""
        if self._updating:
            return
        
        # 表示モードの場合はそのまま実テキストに反映
        if self.toggle_visibility_btn.isChecked():
            self._real_text = displayed_text
            return
        
        self._updating = True
        
        mask_char = '●'
        old_real = self._real_text
        old_len = len(old_real)
        new_len = len(displayed_text)
        
        # カーソル位置を取得
        cursor_pos = self.api_key_input.cursorPosition()
        
        if new_len > old_len:
            # 文字が追加された：マスク文字でない部分が新しい入力
            added_count = new_len - old_len
            insert_pos = cursor_pos - added_count
            new_chars = displayed_text[insert_pos:cursor_pos]
            self._real_text = old_real[:insert_pos] + new_chars + old_real[insert_pos:]
        elif new_len < old_len:
            # 文字が削除された
            deleted_count = old_len - new_len
            self._real_text = old_real[:cursor_pos] + old_real[cursor_pos + deleted_count:]
        else:
            # 長さ同じ（置換操作等）：変更なし
            pass
        
        # マスク表示に更新
        masked = mask_char * len(self._real_text)
        self.api_key_input.setText(masked)
        self.api_key_input.setCursorPosition(cursor_pos)
        
        self._updating = False
    
    def _toggle_visibility(self, checked):
        """APIキーの表示/非表示を切り替え"""
        self._updating = True
        if checked:
            self.toggle_visibility_btn.setText("隠す")
            self.api_key_input.setText(self._real_text)
        else:
            self.toggle_visibility_btn.setText("表示")
            masked = '●' * len(self._real_text)
            self.api_key_input.setText(masked)
        self._updating = False
    
    def get_api_key(self):
        return self._real_text.strip()


class PDFAnalyzer(QThread):
    """PDFを分析してExcelを生成するスレッド"""
    
    progress = pyqtSignal(int, int, str)  # current, total, message
    finished = pyqtSignal(str, str)  # output_path, rename_folder
    error = pyqtSignal(str)  # error_message
    
    def __init__(self, pdf_files, api_key, output_path):
        super().__init__()
        self.pdf_files = pdf_files
        self.api_key = api_key
        self.output_path = output_path
    
    def run(self):
        try:
            # Excelワークブックを作成
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "証拠説明書"

            # ヘッダー行（H列=ファイルパスは非表示）
            headers = ["証拠種別", "番号", "標題", "原本・写しの別", "作成者", "作成年月日", "立証趣旨", "ファイルパス"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = ExcelFont(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 各PDFを処理
            for idx, pdf_info in enumerate(self.pdf_files, 1):
                self.progress.emit(idx, len(self.pdf_files), f"処理中: {pdf_info['filename']}")

                # PDFの1ページ目を画像として取得
                image_base64 = self.extract_first_page_image(pdf_info['path'])

                # Claude APIで分析
                result = self.analyze_with_claude(image_base64)

                # 証拠番号を分割
                evidence_type, evidence_num = self.split_evidence_number(pdf_info['evidence_number'])

                # Excelに書き込み（H列にファイルパスを格納）
                row = idx + 1
                ws.cell(row, 1, evidence_type)
                ws.cell(row, 2, evidence_num)
                ws.cell(row, 3, result.get('title', ''))
                ws.cell(row, 4, '写し')
                ws.cell(row, 5, result.get('author', ''))
                ws.cell(row, 6, to_wareki(result.get('date', '')))
                ws.cell(row, 7, result.get('purpose', ''))
                ws.cell(row, 8, pdf_info['path'])  # ファイルパス（非表示）

            # 列幅を調整
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 40
            # H列（ファイルパス）を非表示
            ws.column_dimensions['H'].hidden = True
            ws.column_dimensions['H'].width = 60

            # 保存（_番号付フォルダに直接保存）
            wb.save(self.output_path)
            self.finished.emit(self.output_path, '')

        except Exception as e:
            self.error.emit(f"エラーが発生しました:\n{str(e)}")
    
    def extract_first_page_image(self, pdf_path):
        """PDFの1ページ目を画像として抽出しbase64エンコード"""
        pdf = fitz.open(pdf_path)
        first_page = pdf[0]
        
        # 高解像度で画像化（zoom=3で精度向上）
        zoom = 3
        mat = fitz.Matrix(zoom, zoom)
        pix = first_page.get_pixmap(matrix=mat)
        
        # PNG形式でバイト列として取得
        img_bytes = pix.tobytes("png")
        
        # base64エンコード
        img_base64 = base64.b64encode(img_bytes).decode('utf-8')
        
        pdf.close()
        return img_base64
    
    def split_evidence_number(self, evidence_number):
        """証拠番号を種別と番号に分割（短縮形式対応）
        例：「甲第1号証」   → （「甲」、「1」）
            「甲第1号証の2」 → （「甲」、「1の2」）
            「第5号証」      → （「」、「5」）
        """
        # パターン1: 甲第1号証の2 など（枝番あり）
        match = re.match(r'(甲|乙|丙|丁)第(.+)号証の(.+)', evidence_number)
        if match:
            return match.group(1), f"{match.group(2)}の{match.group(3)}"

        # パターン2: 甲第1号証 など（枝番なし）
        match = re.match(r'(甲|乙|丙|丁)第(.+)号証', evidence_number)
        if match:
            return match.group(1), match.group(2)

        # パターン3: 第1号証など
        match = re.match(r'第(.+)号証', evidence_number)
        if match:
            return '', match.group(1)

        # マッチしない場合
        return '', evidence_number
    
    def analyze_with_claude(self, image_base64):
        """Claude APIで画像を分析"""
        try:
            import requests
            
            headers = {
                "x-api-key": self.api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json"
            }
            
            prompt = """あなたは法律事務所の証拠整理を支援しています。
この文書は訴訟の証拠として使用され、証拠説明書に記載する情報を抽出する必要があります。

証拠説明書には以下の項目を記載します。この目的を踏まえて、適切な情報を抽出してください。

【重要な注意事項】
- 手書きの文字は無視してください。印刷されたテキストのみを対象としてください。
- スキャン画像で不鮮明な場合でも、読み取れる範囲で最善を尽くしてください。

1. 標題（文書のタイトル）
   - 最も大きなフォントサイズで印刷されているテキスト、または2番目に大きなフォントサイズのテキストが、文書のタイトルであることが多いです
   - タイトルらしい大きな印刷文字列を優先的に抽出してください
   - 会社名や発行者名ではなく、文書の内容を示すタイトルを抽出してください
   - 例：「業務委託契約書」「請求書」「議事録」「覚書」など

2. 作成者（会社名・個人名）
   - 文書の発行者や作成者を抽出してください（印刷されたもののみ）
   - 例：「株式会社○○」「○○法律事務所」「山田太郎」など

3. 作成年月日
   - 文書に記載されている日付を抽出してください（印刷されたもののみ）
   - 形式: YYYY年MM月DD日（例: 2024年3月15日）
   - 複数の日付がある場合は、文書の作成日や発行日と思われる日付を優先してください

4. 立証趣旨
   - 文書の種類と内容から、訴訟で何を立証するために使用される文書かを推測してください
   - 必ず体言止めで記載してください（「〜こと」「〜の事実」などで終わる）
   - 例：
     * 業務委託契約書 → 「業務委託契約の成立」
     * 請求書 → 「請求金額の存在」
     * 領収書 → 「支払いの事実」
     * 議事録 → 「協議の経緯」
     * 覚書 → 「合意内容の確認」
     * メール → 「意思表示の事実」
     * 見積書 → 「見積金額の提示」
   - 簡潔に、10文字程度で記載してください

情報が見つからない場合は空文字列を返してください。
以下のJSON形式で回答してください：

{
  "title": "文書のタイトル",
  "author": "作成者名",
  "date": "YYYY年MM月DD日",
  "purpose": "立証趣旨（体言止め）"
}"""
            
            data = {
                "model": "claude-opus-4-6",
                "max_tokens": 1000,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/png",
                                    "data": image_base64
                                }
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ]
            }
            
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers,
                json=data,
                timeout=30
            )
            
            if response.status_code == 200:
                result = response.json()
                text = result['content'][0]['text']
                
                # JSONを抽出（```json ... ```の場合も対応）
                text = text.strip()
                if text.startswith('```json'):
                    text = text[7:]
                if text.startswith('```'):
                    text = text[3:]
                if text.endswith('```'):
                    text = text[:-3]
                text = text.strip()
                
                return json.loads(text)
            else:
                return {"title": "", "author": "", "date": "", "purpose": ""}
                
        except Exception as e:
            print(f"API Error: {e}")
            return {"title": "", "author": "", "date": "", "purpose": ""}


class EvidenceScheduleWindow(QMainWindow):
    """証拠説明書自動作成ツール メインウィンドウ"""
    
    def __init__(self):
        super().__init__()
        self.api_key = self.load_api_key()
        self.pdf_files = []
        self.init_ui()
    
    def init_ui(self):
        """UIの初期化"""
        self.setWindowTitle("証拠説明書作成支援ツール")
        self.setMinimumSize(800, 600)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # ヘッダー
        header_layout = QHBoxLayout()
        
        title_label = QLabel("証拠説明書作成支援ツール")
        title_label.setFont(QFont("メイリオ", 14, QFont.Weight.Bold))
        
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        
        # Aboutリンク
        about_label = QLabel('<a href="#" style="color: #888;">About</a>')
        about_label.setOpenExternalLinks(False)
        about_label.linkActivated.connect(lambda: show_about_dialog())
        header_layout.addWidget(about_label)
        
        main_layout.addLayout(header_layout)
        
        # 使い方説明
        help_label = QLabel(
            "使い方：\n"
            "1. 「フォルダを開く」でPDFファイルを読み込む\n"
            "2. 「生成実行」でClaude APIを使って証拠説明書案を生成"
        )
        help_label.setStyleSheet("""
            QLabel {
                background-color: #FFF9C4;
                padding: 10px;
                border: 1px solid #FBC02D;
            }
        """)
        main_layout.addWidget(help_label)
        
        # ボタンエリア
        button_layout = QHBoxLayout()
        
        self.load_folder_btn = QPushButton("フォルダを開く")
        self.load_folder_btn.clicked.connect(self.load_folder)
        
        self.add_files_btn = QPushButton("ファイルを追加")
        self.add_files_btn.clicked.connect(self.add_files)
        
        self.clear_btn = QPushButton("クリア")
        self.clear_btn.clicked.connect(self.clear_list)
        
        self.api_key_btn = QPushButton("APIキー設定")
        self.api_key_btn.clicked.connect(self.show_api_key_dialog)
        
        self.delete_api_key_btn = QPushButton("APIキー削除")
        self.delete_api_key_btn.clicked.connect(self.delete_api_key)
        
        button_layout.addWidget(self.load_folder_btn)
        button_layout.addWidget(self.add_files_btn)
        button_layout.addWidget(self.clear_btn)
        button_layout.addWidget(self.api_key_btn)
        button_layout.addWidget(self.delete_api_key_btn)
        button_layout.addStretch()
        
        main_layout.addLayout(button_layout)
        
        # ファイルリスト
        list_label = QLabel("読み込んだPDFファイル:")
        main_layout.addWidget(list_label)
        
        self.file_list = QListWidget()
        main_layout.addWidget(self.file_list)
        
        # プログレスバー
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        # 実行ボタン
        execute_layout = QHBoxLayout()

        self.execute_btn = QPushButton("生成実行")
        self.execute_btn.clicked.connect(self.execute_generation)
        self.execute_btn.setEnabled(False)
        self.execute_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 12pt;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #CCCCCC;
            }
        """)

        execute_layout.addStretch()
        execute_layout.addWidget(self.execute_btn)

        main_layout.addLayout(execute_layout)

        # ─── リネーム実行エリア ───
        rename_group_label = QLabel("─── PDFリネーム（任意）───")
        rename_group_label.setStyleSheet("color: #888; margin-top: 10px;")
        main_layout.addWidget(rename_group_label)

        rename_help = QLabel(
            "使い方：\n"
            "証拠説明書案のExcel/CSVの内容確認・修正後にドロップ → リネーム実行\n"
            "対応する全PDFのファイル名末尾に証拠標題が付加され、「_リネーム済」フォルダに格納されます"
        )
        rename_help.setStyleSheet("""
            QLabel {
                background-color: #F0F4FF;
                padding: 8px;
                border: 1px solid #B0BEC5;
                color: #555;
            }
        """)
        main_layout.addWidget(rename_help)

        # ドロップエリア
        self.drop_label = QLabel("Excel / CSV をここにドロップ")
        self.drop_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.drop_label.setFixedHeight(50)
        self.drop_label.setStyleSheet("""
            QLabel {
                border: 2px dashed #90A4AE;
                background-color: #FAFAFA;
                color: #90A4AE;
                font-size: 11pt;
            }
        """)
        self.drop_label.setAcceptDrops(True)
        self.drop_label.dragEnterEvent = self._drag_enter
        self.drop_label.dropEvent = self._drop_file
        main_layout.addWidget(self.drop_label)

        # ボタン行（ファイルをクリア：左、リネーム実行：右）
        rename_btn_layout = QHBoxLayout()

        self.clear_rename_btn = QPushButton("ファイルをクリア")
        self.clear_rename_btn.setEnabled(False)
        self.clear_rename_btn.clicked.connect(self._clear_rename_file)

        self.rename_btn = QPushButton("リネーム実行")
        self.rename_btn.setEnabled(False)
        self.rename_btn.clicked.connect(self.execute_rename)
        self.rename_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                font-size: 12pt;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #1565C0;
            }
            QPushButton:disabled {
                background-color: #CCCCCC;
            }
        """)

        rename_btn_layout.addWidget(self.clear_rename_btn)
        rename_btn_layout.addStretch()
        rename_btn_layout.addWidget(self.rename_btn)
        main_layout.addLayout(rename_btn_layout)

        # 内部状態
        self._rename_table_path = None  # ドロップされたExcel/CSVパス
        
        # APIキーチェック
        if not self.api_key:
            QMessageBox.information(
                self,
                "APIキー設定",
                "初回起動時にAPIキーの設定が必要です。"
            )
            self.show_api_key_dialog()
    
    def load_api_key(self):
        """APIキーを読み込み（セキュアストレージから）"""
        return load_api_key_secure()
    
    def save_api_key(self, api_key):
        """APIキーを保存（セキュアストレージへ）"""
        save_api_key_secure(api_key)
    
    def show_api_key_dialog(self):
        """APIキー設定ダイアログを表示"""
        dialog = APIKeyDialog(self)
        if dialog.exec():
            api_key = dialog.get_api_key()
            if api_key:
                try:
                    self.api_key = api_key
                    self.save_api_key(api_key)
                    QMessageBox.information(
                        self, 
                        "設定完了", 
                        f"APIキーを保存しました。\n使用モデル: Opus 4.6（最高精度）"
                    )
                except RuntimeError as e:
                    # 保存用ライブラリがない場合：メモリ上のみで保持
                    self.api_key = api_key
                    QMessageBox.warning(
                        self,
                        "注意",
                        f"{str(e)}\n\nAPIキーは今回の起動中のみ有効です。"
                    )
    
    def load_folder(self):
        """フォルダからPDFファイルを読み込み"""
        folder_path = QFileDialog.getExistingDirectory(self, "フォルダを選択")
        if not folder_path:
            return
        
        pdf_files = list(Path(folder_path).glob("*.pdf"))
        
        if not pdf_files:
            QMessageBox.warning(self, "警告", "PDFファイルが見つかりません。")
            return
        
        # ファイル名から証拠番号を抽出してソート
        self.pdf_files = self.sort_and_number_files(pdf_files)
        
        # リストに表示
        self.update_file_list()
        
        self.execute_btn.setEnabled(bool(self.pdf_files))
    
    def add_files(self):
        """個別ファイルを追加"""
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "PDFファイルを選択",
            "",
            "PDF Files (*.pdf)"
        )
        
        if not file_paths:
            return
        
        # 既存のファイルと新しいファイルを結合
        existing_paths = [item['path'] for item in self.pdf_files]
        all_paths = existing_paths + [Path(p) for p in file_paths if p not in existing_paths]
        
        # 再ソート
        self.pdf_files = self.sort_and_number_files(all_paths)
        
        # リストに表示
        self.update_file_list()
        
        self.execute_btn.setEnabled(bool(self.pdf_files))
    
    def update_file_list(self):
        """ファイルリストを更新"""
        self.file_list.clear()
        for pdf_info in self.pdf_files:
            short = format_evidence_short(pdf_info['evidence_number'])
            item = QListWidgetItem(f"{short} - {pdf_info['filename']}")
            self.file_list.addItem(item)
    
    def delete_api_key(self):
        """APIキーを削除"""
        reply = QMessageBox.question(
            self,
            "確認",
            "保存されているAPIキーを削除しますか？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            delete_api_key_secure()
            self.api_key = ''
            QMessageBox.information(self, "完了", "APIキーを削除しました。")
    
    def sort_and_number_files(self, pdf_files):
        """ファイルをソートして証拠番号を付与"""
        categorized = {
            '甲': [],
            '乙': [],
            '丙': [],
            '丁': [],
            'other': []
        }
        
        # ファイルをカテゴリ分け
        for pdf_path in pdf_files:
            filename = pdf_path.stem

            # 証拠種別と番号を抽出
            # 対応形式: 甲01、甲01の1、甲01の１（全角）、甲01-1
            match = re.search(r'(甲|乙|丙|丁)?(\d+)(?:[のの](\d+)|[-－](\d+))?', filename)

            if match:
                category = match.group(1) if match.group(1) else 'other'
                main_num = match.group(2)
                # 「の」区切り（全角・半角）またはハイフン区切りの枝番
                branch_num = match.group(3) or match.group(4)
                number = f"{main_num}-{branch_num}" if branch_num else main_num

                if category in categorized:
                    categorized[category].append({
                        'path': str(pdf_path),
                        'filename': pdf_path.name,
                        'number': number,
                        'sort_key': self.parse_number(number)
                    })
                else:
                    categorized['other'].append({
                        'path': str(pdf_path),
                        'filename': pdf_path.name,
                        'number': None,
                        'sort_key': (999999, 0)
                    })
            else:
                categorized['other'].append({
                    'path': str(pdf_path),
                    'filename': pdf_path.name,
                    'number': None,
                    'sort_key': (999999, 0)
                })
        
        # 各カテゴリ内でソート
        for key in categorized:
            categorized[key].sort(key=lambda x: x['sort_key'])
        
        # 結合して証拠番号を付与
        result = []
        counter = 1
        
        for category in ['甲', '乙', '丙', '丁']:
            for item in categorized[category]:
                # ゼロ埋め除去・枝番を「の」形式に変換
                # 例: "01" → "1"、"01-2" → "1の2"
                raw = item['number']
                if '-' in raw:
                    parts = raw.split('-', 1)
                    main = int(parts[0])
                    branch = int(parts[1])
                    evidence_number = f"{category}第{main}号証の{branch}"
                else:
                    main = int(raw)
                    evidence_number = f"{category}第{main}号証"
                result.append({
                    'path': item['path'],
                    'filename': item['filename'],
                    'evidence_number': evidence_number
                })
        
        # 番号なしファイルは自動採番
        for item in categorized['other']:
            # すでに番号が振られているものをスキップして次の番号を決定
            while any(f"第{counter}号証" in r['evidence_number'] for r in result):
                counter += 1
            
            evidence_number = f"第{counter}号証"
            result.append({
                'path': item['path'],
                'filename': item['filename'],
                'evidence_number': evidence_number
            })
            counter += 1
        
        return result
    
    def parse_number(self, number_str):
        """番号文字列をソート用のタプルに変換"""
        if '-' in number_str:
            parts = number_str.split('-')
            return (int(parts[0]), int(parts[1]))
        else:
            return (int(number_str), 0)
    
    def clear_list(self):
        """リストをクリア"""
        self.pdf_files = []
        self.file_list.clear()
        self.execute_btn.setEnabled(False)
    
    def execute_generation(self):
        """証拠説明書生成を実行"""
        if not self.api_key:
            QMessageBox.warning(self, "警告", "APIキーが設定されていません。")
            self.show_api_key_dialog()
            return

        if not self.pdf_files:
            QMessageBox.warning(self, "警告", "PDFファイルが読み込まれていません。")
            return

        # 保存先フォルダ：最初のファイルの親フォルダを自動使用
        first_file_path = Path(self.pdf_files[0]['path'])
        output_folder = first_file_path.parent

        # ファイル名：フォルダ名_証拠説明書案.xlsx
        folder_name = output_folder.name
        filename = f"{folder_name}_証拠説明書案.xlsx"
        output_path = output_folder / filename

        # リネーム済フォルダのパスを予測（スレッド内と同じロジック）
        rename_folder_preview = output_folder.parent / (output_folder.name + "_リネーム済")

        # 同名ファイルが存在する場合は上書き確認
        excel_in_rename = rename_folder_preview / filename
        if excel_in_rename.exists():
            overwrite_reply = QMessageBox.question(
                self,
                "上書き確認",
                f"既に同名のファイルが存在します。\n{excel_in_rename}\n\n上書きしますか？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if overwrite_reply != QMessageBox.StandardButton.Yes:
                return

        # 実行確認ダイアログ
        reply = QMessageBox.question(
            self,
            "確認",
            f"{len(self.pdf_files)}件のPDFを処理します。\n"
            f"API料金が発生します（概算: ${len(self.pdf_files) * 0.01:.2f}）。\n"
            f"保存先: {rename_folder_preview}\n\n"
            "実行しますか？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply != QMessageBox.StandardButton.Yes:
            return

        # プログレスバーを表示
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(len(self.pdf_files))
        self.progress_bar.setValue(0)

        # ボタンを無効化
        self.execute_btn.setEnabled(False)
        self.load_folder_btn.setEnabled(False)

        # バックグラウンドで処理
        self.analyzer = PDFAnalyzer(self.pdf_files, self.api_key, str(output_path))
        self.analyzer.progress.connect(self.on_progress)
        self.analyzer.finished.connect(self.on_finished)
        self.analyzer.error.connect(self.on_error)
        self.analyzer.start()
    
    def on_progress(self, current, total, message):
        """進捗更新"""
        self.progress_bar.setValue(current)
        self.statusBar().showMessage(message)
    
    def on_finished(self, output_path, _rename_folder):
        """処理完了"""
        self.progress_bar.setVisible(False)
        self.execute_btn.setEnabled(True)
        self.load_folder_btn.setEnabled(True)
        self.statusBar().showMessage("完了")

        QMessageBox.information(
            self,
            "完了",
            f"証拠説明書を作成しました:\n{output_path}"
        )

        # 出力フォルダを開く
        import subprocess
        import platform
        if platform.system() == 'Windows':
            subprocess.Popen(['start', '', str(Path(output_path).parent)], shell=True)
    
    def on_error(self, error_message):
        """エラー発生"""
        self.progress_bar.setVisible(False)
        self.execute_btn.setEnabled(True)
        self.load_folder_btn.setEnabled(True)
        self.statusBar().showMessage("エラー")
        QMessageBox.critical(self, "エラー", error_message)

    # ─── リネーム機能 ───────────────────────────────────────

    def _drag_enter(self, event):
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0].toLocalFile()
            if url.lower().endswith(('.xlsx', '.csv')):
                event.acceptProposedAction()
                return
        event.ignore()

    def _drop_file(self, event):
        urls = event.mimeData().urls()
        if not urls:
            return
        file_path = urls[0].toLocalFile()
        if not file_path.lower().endswith(('.xlsx', '.csv')):
            QMessageBox.warning(self, "警告", "xlsx または csv ファイルをドロップしてください。")
            return
        self._rename_table_path = file_path
        self.drop_label.setText(f"✓ {Path(file_path).name}")
        self.drop_label.setStyleSheet("""
            QLabel {
                border: 2px solid #4CAF50;
                background-color: #F1F8E9;
                color: #2E7D32;
                font-size: 11pt;
            }
        """)
        self.rename_btn.setEnabled(True)

    def _clear_rename_file(self):
        """ドロップファイルをクリア"""
        self._rename_table_path = None
        self.drop_label.setText("Excel / CSV をここにドロップ")
        self.drop_label.setStyleSheet("""
            QLabel {
                border: 2px dashed #90A4AE;
                background-color: #FAFAFA;
                color: #90A4AE;
                font-size: 11pt;
            }
        """)
        self.rename_btn.setEnabled(False)
        self.clear_rename_btn.setEnabled(False)

    def execute_rename(self):
        import shutil, csv
        if not self._rename_table_path:
            return

        # Excel/CSVを読み込む
        table_path = Path(self._rename_table_path)
        rows = []
        try:
            if table_path.suffix.lower() == '.xlsx':
                wb_r = openpyxl.load_workbook(str(table_path))
                ws_r = wb_r.active
                for row in ws_r.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    etype = str(row[0]).strip() if row[0] else ''
                    enum  = str(row[1]).strip() if row[1] else ''
                    title = str(row[2]).strip() if row[2] else ''
                    fp = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                    rows.append((fp, etype, enum, title))
            else:
                with open(str(table_path), encoding='utf-8-sig', newline='') as f:
                    reader = csv.reader(f)
                    next(reader, None)
                    for row in reader:
                        if not any(row):
                            continue
                        etype = row[0].strip() if len(row) > 0 else ''
                        enum  = row[1].strip() if len(row) > 1 else ''
                        title = row[2].strip() if len(row) > 2 else ''
                        fp = row[7].strip() if len(row) > 7 else ''
                        rows.append((fp, etype, enum, title))
        except Exception as e:
            QMessageBox.critical(self, "エラー", f"ファイルの読み込みに失敗しました:\n{str(e)}")
            return

        if not rows:
            QMessageBox.warning(self, "警告", "データが見つかりませんでした。")
            return

        # H列パスの生死チェック：1件でも生きていればそのまま処理
        # 全件パスが死んでいる場合のみフォルダ選択を求める
        pdf_folder = None
        all_paths_dead = all(not Path(fp).exists() for fp, *_ in rows if fp)

        if all_paths_dead:
            answer = QMessageBox.warning(
                self, "ファイルパスが見つかりません",
                "ファイルパスが見つかりません。\n_番号付フォルダを選択してください。",
                QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel
            )
            if answer != QMessageBox.StandardButton.Ok:
                return
            folder = QFileDialog.getExistingDirectory(
                self, "_番号付フォルダを選択してください"
            )
            if not folder:
                return
            pdf_folder = Path(folder)

        # リネーム済フォルダを作成
        # 基準フォルダ：H列パスから、またはユーザー指定フォルダから決定
        if pdf_folder:
            base_folder = pdf_folder
        else:
            # H列の最初の生きているパスの親フォルダを基準にする
            base_folder = next(
                (Path(fp).parent for fp, *_ in rows if fp and Path(fp).exists()),
                table_path.parent
            )

        rename_folder_base = base_folder.parent / (base_folder.name + "_リネーム済")
        rename_folder = rename_folder_base
        suffix = 2
        while rename_folder.exists():
            rename_folder = rename_folder_base.parent / (rename_folder_base.name + f"_{suffix}")
            suffix += 1
        rename_folder.mkdir(parents=True, exist_ok=True)

        success, skipped = 0, 0
        for fp_str, etype, enum, title in rows:
            if fp_str:
                src = Path(fp_str)
                # H列パスが死んでいる場合はフォルダ指定から探す
                if not src.exists() and pdf_folder:
                    src = pdf_folder / src.name
            else:
                skipped += 1
                continue
            if not src.exists():
                skipped += 1
                continue
            # ファイル名を「証拠種別＋番号（ゼロ埋め）＋標題」で生成
            # 例: 甲+1+契約書 → 甲01_契約書.pdf
            if etype and enum:
                try:
                    # 枝番対応: "1の2" → "01の2"
                    if 'の' in enum:
                        main, branch = enum.split('の', 1)
                        num_prefix = f"{etype}{int(main):02d}の{branch}"
                    else:
                        num_prefix = f"{etype}{int(enum):02d}"
                except ValueError:
                    num_prefix = f"{etype}{enum}"
            else:
                # 種別・番号が空欄の場合は元ファイル名をそのまま使う
                num_prefix = src.stem
            if title:
                safe_title = re.sub(r'[\\/:*?"<>|]', '', title)
                new_name = f"{num_prefix}_{safe_title}.pdf"
            else:
                new_name = f"{num_prefix}.pdf"
            shutil.copy2(str(src), str(rename_folder / new_name))
            success += 1

        msg = f"リネーム完了：{success}件\n出力先：{rename_folder}"
        if skipped:
            msg += f"\n\nスキップ（ファイル未発見）：{skipped}件"
        QMessageBox.information(self, "完了", msg)

        # 完了後に自動クリア
        self._clear_rename_file()

        import subprocess, platform
        if platform.system() == 'Windows':
            subprocess.Popen(['start', '', str(rename_folder)], shell=True)


def main():
    app = QApplication(sys.argv)
    window = EvidenceScheduleWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
